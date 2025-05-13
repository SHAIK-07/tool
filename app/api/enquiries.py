from fastapi import APIRouter, Depends, Form, Request, Body, HTTPException, Query, File, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, FileResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from typing import Optional, List, Annotated
from datetime import datetime, date, timedelta
import json
import os
import csv
import shutil
from io import StringIO

from app.db import crud, database, models
from app.schemas import enquiry as enquiry_schemas
from app.core.auth import get_current_user_from_cookie  # Import once at the top

# Define quotations folder path
QUOTATIONS_FOLDER = "quotations"
os.makedirs(QUOTATIONS_FOLDER, exist_ok=True)

router = APIRouter(tags=["Enquiries"])
templates = Jinja2Templates(directory="templates")


@router.get("/enquiries", response_class=HTMLResponse)
async def get_enquiries_page(
    request: Request,
    message: str = None,
    page: int = 1,
    limit: int = 50,
    customer_name: str = None,
    date_from: str = None,
    date_to: str = None,
    quotation_given: bool = None,
    db: Session = Depends(database.get_db)
):
    """Display the enquiries page with a list of all enquiries and the form to add new enquiries"""
    try:
        # Get current user from cookie
        user = await get_current_user_from_cookie(request, db)

        # Redirect to login if not authenticated
        if not user:
            return RedirectResponse(url="/login?next=/enquiries", status_code=303)

        # Apply filters
        filter_params = {}
        if customer_name:
            filter_params["customer_name"] = customer_name

        if date_from:
            try:
                filter_params["date_from"] = datetime.strptime(date_from, "%Y-%m-%d").date()
            except ValueError:
                # Invalid date format, ignore this filter
                pass

        if date_to:
            try:
                filter_params["date_to"] = datetime.strptime(date_to, "%Y-%m-%d").date()
            except ValueError:
                # Invalid date format, ignore this filter
                pass

        if quotation_given is not None:
            filter_params["quotation_given"] = quotation_given

        # Calculate offset for pagination
        offset = (page - 1) * limit

        # Get total count for pagination
        total_enquiries = crud.get_total_enquiries_count(db)

        # Calculate total pages
        total_pages = (total_enquiries + limit - 1) // limit if total_enquiries > 0 else 1

        # Ensure page is within valid range
        if page < 1:
            page = 1
        elif page > total_pages:
            page = total_pages

        # Recalculate offset after page validation
        offset = (page - 1) * limit

        # Get paginated enquiries from database
        if filter_params:
            # If filters are applied, use filtered query
            enquiries = crud.get_filtered_enquiries(db, filter_params, limit=limit, offset=offset)
        else:
            # Otherwise get all enquiries with pagination
            enquiries = crud.get_paginated_enquiries(db, limit=limit, offset=offset)

        # Get today's date for the form
        today = datetime.now().date()

        return templates.TemplateResponse(
            "enquiries.html",
            {
                "request": request,
                "user": user,
                "enquiries": enquiries,
                "today": today,
                "message": message,
                "current_page": page,
                "total_pages": total_pages,
                "limit": limit,
                "total_items": total_enquiries
            }
        )
    except Exception as e:
        print(f"Error getting enquiries page: {e}")
        return templates.TemplateResponse(
            "error.html",
            {"request": request, "error_message": f"Error loading enquiries: {str(e)}"}
        )


@router.get("/enquiry/new", response_class=HTMLResponse)
async def new_enquiry_form(
    request: Request,
    db: Session = Depends(database.get_db)
):
    """Display the form to create a new enquiry"""
    try:
        # Get current user from cookie
        user = await get_current_user_from_cookie(request, db)

        # Redirect to login if not authenticated
        if not user:
            return RedirectResponse(url="/login?next=/enquiry/new", status_code=303)

        # Pass today's date to the template
        today = datetime.now().date()

        return templates.TemplateResponse(
            "enquiry_form.html",
            {
                "request": request,
                "user": user,
                "edit_mode": False,
                "today": today,
                "enquiry": {}
            }
        )
    except Exception as e:
        print(f"Error displaying new enquiry form: {e}")
        return templates.TemplateResponse(
            "error.html",
            {"request": request, "error_message": f"Error loading form: {str(e)}"}
        )


@router.get("/enquiry/{enquiry_id}/edit", response_class=HTMLResponse)
async def edit_enquiry_form(
    request: Request,
    enquiry_id: int,
    db: Session = Depends(database.get_db)
):
    """Display the form to edit an existing enquiry"""
    try:
        # Get current user from cookie
        user = await get_current_user_from_cookie(request, db)

        # Redirect to login if not authenticated
        if not user:
            return RedirectResponse(url=f"/login?next=/enquiry/{enquiry_id}/edit", status_code=303)

        # Get enquiry from database
        enquiry = crud.get_enquiry(db, enquiry_id)
        if not enquiry:
            return templates.TemplateResponse(
                "error.html",
                {"request": request, "error_message": "Enquiry not found", "user": user},
                status_code=404
            )

        return templates.TemplateResponse(
            "enquiry_form.html",
            {
                "request": request,
                "user": user,
                "edit_mode": True,
                "enquiry": enquiry
            }
        )
    except Exception as e:
        print(f"Error displaying edit enquiry form: {e}")
        return templates.TemplateResponse(
            "error.html",
            {"request": request, "error_message": f"Error loading form: {str(e)}"}
        )


@router.post("/enquiry/new", response_class=HTMLResponse)
async def create_new_enquiry(
    request: Request,
    customer_name: str = Form(...),
    phone_no: str = Form(...),
    address: str = Form(None),
    requirements: str = Form(None),
    quotation_given: bool = Form(False),
    quotation_amount: Optional[float] = Form(0.0),
    date: str = Form(...),
    quotation_file: UploadFile = File(None),
    db: Session = Depends(database.get_db),
):
    """Create a new enquiry record"""
    try:
        # Get current user from cookie
        user = await get_current_user_from_cookie(request, db)

        # Redirect to login if not authenticated
        if not user:
            return RedirectResponse(url="/login?next=/enquiry/new", status_code=303)

        # Convert date string to date object
        try:
            date_obj = datetime.strptime(date, "%Y-%m-%d").date()
        except ValueError:
            return templates.TemplateResponse(
                "enquiry_form.html",
                {
                    "request": request,
                    "user": user,
                    "edit_mode": False,
                    "error": "Invalid date format",
                    "enquiry": {
                        "customer_name": customer_name,
                        "phone_no": phone_no,
                        "address": address,
                        "requirements": requirements,
                        "quotation_given": quotation_given,
                        "quotation_amount": quotation_amount
                    }
                }
            )

        # Handle quotation file upload if quotation is given
        quotation_file_path = None
        if quotation_given and quotation_file and quotation_file.filename:
            # Create a unique filename
            file_extension = quotation_file.filename.split('.')[-1]
            unique_filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{quotation_file.filename}"
            file_path = os.path.join(QUOTATIONS_FOLDER, unique_filename)

            # Save the file
            with open(file_path, "wb") as buffer:
                content = await quotation_file.read()
                buffer.write(content)

            quotation_file_path = unique_filename

        # Create enquiry data dictionary
        enquiry_data = {
            "customer_name": customer_name,
            "phone_no": phone_no,
            "address": address or "",  # Ensure address is not None
            "requirements": requirements or "",  # Ensure requirements is not None
            "quotation_given": quotation_given,
            "quotation_amount": float(quotation_amount) if quotation_given else None,
            "quotation_file_path": quotation_file_path,
            "date": date_obj
        }

        # Create enquiry in database
        enquiry = crud.create_enquiry(db, enquiry_data)

        # Redirect to enquiries page with success message
        return RedirectResponse(url=f"/enquiries?message=Enquiry+created+successfully", status_code=303)
    except Exception as e:
        print(f"Error creating enquiry: {e}")
        return templates.TemplateResponse(
            "enquiry_form.html",
            {
                "request": request,
                "user": user,
                "edit_mode": False,
                "error": f"Error creating enquiry: {str(e)}",
                "enquiry": enquiry_data
            }
        )


@router.post("/enquiry/{enquiry_id}/edit", response_class=HTMLResponse)
async def update_enquiry(
    request: Request,
    enquiry_id: int,
    customer_name: str = Form(...),
    phone_no: str = Form(...),
    address: str = Form(None),
    requirements: str = Form(None),
    quotation_given: bool = Form(False),
    quotation_amount: Optional[float] = Form(0.0),
    date: str = Form(...),
    quotation_file: UploadFile = File(None),
    db: Session = Depends(database.get_db)
):
    """Update an enquiry record"""
    try:
        # Get current user from cookie
        user = await get_current_user_from_cookie(request, db)

        # Redirect to login if not authenticated
        if not user:
            return RedirectResponse(url=f"/login?next=/enquiry/{enquiry_id}/edit", status_code=303)

        # Get existing enquiry to check if we need to update the file
        existing_enquiry = crud.get_enquiry(db, enquiry_id)
        if not existing_enquiry:
            return templates.TemplateResponse(
                "error.html",
                {"request": request, "error_message": "Enquiry not found", "user": user},
                status_code=404
            )

        # Convert date string to date object
        try:
            date_obj = datetime.strptime(date, "%Y-%m-%d").date()
        except ValueError:
            return templates.TemplateResponse(
                "enquiry_form.html",
                {
                    "request": request,
                    "user": user,
                    "edit_mode": True,
                    "error": "Invalid date format",
                    "enquiry": {
                        "id": enquiry_id,
                        "customer_name": customer_name,
                        "phone_no": phone_no,
                        "address": address,
                        "requirements": requirements,
                        "quotation_given": quotation_given,
                        "quotation_amount": quotation_amount,
                        "quotation_file_path": existing_enquiry.quotation_file_path
                    }
                }
            )

        # Handle quotation file upload if quotation is given
        quotation_file_path = existing_enquiry.quotation_file_path
        if quotation_given and quotation_file and quotation_file.filename:
            # Create a unique filename
            file_extension = quotation_file.filename.split('.')[-1]
            unique_filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{quotation_file.filename}"
            file_path = os.path.join(QUOTATIONS_FOLDER, unique_filename)

            # Save the file
            with open(file_path, "wb") as buffer:
                content = await quotation_file.read()
                buffer.write(content)

            # Delete old file if it exists
            if existing_enquiry.quotation_file_path:
                old_file_path = os.path.join(QUOTATIONS_FOLDER, existing_enquiry.quotation_file_path)
                if os.path.exists(old_file_path):
                    try:
                        os.remove(old_file_path)
                    except Exception as e:
                        print(f"Error deleting old file: {e}")

            quotation_file_path = unique_filename
        elif not quotation_given and existing_enquiry.quotation_file_path:
            # If quotation is no longer given, delete the file
            old_file_path = os.path.join(QUOTATIONS_FOLDER, existing_enquiry.quotation_file_path)
            if os.path.exists(old_file_path):
                try:
                    os.remove(old_file_path)
                except Exception as e:
                    print(f"Error deleting old file: {e}")
            quotation_file_path = None

        # Create enquiry data dictionary
        enquiry_data = {
            "customer_name": customer_name,
            "phone_no": phone_no,
            "address": address or "",  # Ensure address is not None
            "requirements": requirements or "",  # Ensure requirements is not None
            "quotation_given": quotation_given,
            "quotation_amount": float(quotation_amount) if quotation_given else None,
            "quotation_file_path": quotation_file_path,
            "date": date_obj
        }

        # Update enquiry in database
        enquiry = crud.update_enquiry(db, enquiry_id, enquiry_data)
        if not enquiry:
            return templates.TemplateResponse(
                "error.html",
                {"request": request, "error_message": "Enquiry not found", "user": user},
                status_code=404
            )

        # Redirect to enquiries page with success message
        return RedirectResponse(url=f"/enquiries?message=Enquiry+updated+successfully", status_code=303)
    except Exception as e:
        print(f"Error updating enquiry: {e}")
        return templates.TemplateResponse(
            "enquiry_form.html",
            {
                "request": request,
                "user": user,
                "edit_mode": True,
                "error": f"Error updating enquiry: {str(e)}",
                "enquiry": enquiry_data
            }
        )


@router.get("/api/enquiries")
async def api_get_enquiries(
    db: Session = Depends(database.get_db)
):
    """Get all enquiries as JSON"""
    enquiries = crud.get_all_enquiries(db)

    result = []
    for e in enquiries:
        result.append({
            "id": e.id,
            "enquiry_number": e.enquiry_number,
            "date": e.date.isoformat(),
            "customer_name": e.customer_name,
            "phone_no": e.phone_no,
            "address": e.address,
            "requirements": e.requirements,
            "quotation_given": e.quotation_given,
            "quotation_amount": e.quotation_amount,
            "created_at": e.created_at.isoformat(),
            "updated_at": e.updated_at.isoformat()
        })

    return result


@router.delete("/api/enquiry/{enquiry_id}")
async def api_delete_enquiry(
    enquiry_id: int,
    db: Session = Depends(database.get_db)
):
    """Delete an enquiry"""
    success = crud.delete_enquiry(db, enquiry_id)
    if not success:
        raise HTTPException(status_code=404, detail="Enquiry not found")
    return {"success": True}


@router.post("/api/enquiry/{enquiry_id}/delete")
async def api_delete_enquiry_post(
    request: Request,
    enquiry_id: int,
    db: Session = Depends(database.get_db)
):
    """Delete an enquiry (POST method)"""
    try:
        # Get current user from cookie
        user = await get_current_user_from_cookie(request, db)

        # Redirect to login if not authenticated
        if not user:
            return RedirectResponse(url="/login?next=/enquiries", status_code=303)

        # Delete the enquiry
        success = crud.delete_enquiry(db, enquiry_id)
        if not success:
            return RedirectResponse(url="/enquiries?error=Enquiry+not+found", status_code=303)

        # Redirect to enquiries page with success message
        return RedirectResponse(url="/enquiries?message=Enquiry+deleted+successfully", status_code=303)
    except Exception as e:
        print(f"Error deleting enquiry: {e}")
        return RedirectResponse(url=f"/enquiries?error=Error+deleting+enquiry:+{str(e)}", status_code=303)


@router.get("/enquiry/{enquiry_id}/delete")
async def delete_enquiry_get(
    request: Request,
    enquiry_id: int,
    db: Session = Depends(database.get_db)
):
    """Delete an enquiry (GET method for direct browser navigation)"""
    try:
        # Get current user from cookie
        user = await get_current_user_from_cookie(request, db)

        # Redirect to login if not authenticated
        if not user:
            return RedirectResponse(url="/login?next=/enquiries", status_code=303)

        # Get the enquiry to check if it exists
        enquiry = crud.get_enquiry(db, enquiry_id)
        if not enquiry:
            return RedirectResponse(url="/enquiries?error=Enquiry+not+found", status_code=303)

        # Delete the enquiry
        success = crud.delete_enquiry(db, enquiry_id)
        if not success:
            return RedirectResponse(url="/enquiries?error=Error+deleting+enquiry", status_code=303)

        # Redirect to enquiries page with success message
        return RedirectResponse(url="/enquiries?message=Enquiry+deleted+successfully", status_code=303)
    except Exception as e:
        print(f"Error deleting enquiry: {e}")
        return RedirectResponse(url=f"/enquiries?error=Error+deleting+enquiry:+{str(e)}", status_code=303)


@router.get("/delete-test", response_class=HTMLResponse)
async def delete_test_page(
    request: Request,
    db: Session = Depends(database.get_db)
):
    """Display the delete test page"""
    try:
        # Get current user from cookie
        user = await get_current_user_from_cookie(request, db)

        # Redirect to login if not authenticated
        if not user:
            return RedirectResponse(url="/login?next=/delete-test", status_code=303)

        return templates.TemplateResponse(
            "delete_test.html",
            {
                "request": request,
                "user": user
            }
        )
    except Exception as e:
        print(f"Error displaying delete test page: {e}")
        return templates.TemplateResponse(
            "error.html",
            {"request": request, "error_message": f"Error loading delete test page: {str(e)}"}
        )


@router.get("/api/enquiries/export-excel")
async def export_enquiries_excel(
    db: Session = Depends(database.get_db)
):
    """Export enquiries to Excel"""
    try:
        # Get all enquiries from database
        enquiries = crud.get_all_enquiries(db)

        # Use the excel_generator module to create the Excel file
        from app.core.excel_generator import export_enquiries_to_excel
        filename, file_path = export_enquiries_to_excel(enquiries)

        # Return the Excel file
        return FileResponse(
            file_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename
        )
    except Exception as e:
        print(f"Error exporting enquiries to Excel: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.post("/api/enquiries/import-excel")
async def import_enquiries_excel(
    file: UploadFile = File(...),
    db: Session = Depends(database.get_db)
):
    """Import enquiries from Excel file"""
    try:
        import pandas as pd
        import tempfile
        import os
        from datetime import datetime

        # Create a temporary file to store the uploaded Excel
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            # Write the uploaded file content to the temporary file
            content = await file.read()
            temp_file.write(content)
            temp_file_path = temp_file.name

        # Read the Excel or CSV file
        try:
            file_extension = file.filename.split('.')[-1].lower()
            if file_extension == 'csv':
                df = pd.read_csv(temp_file_path)
            else:
                df = pd.read_excel(temp_file_path)
        except Exception as e:
            os.unlink(temp_file_path)  # Delete the temp file
            return JSONResponse(
                status_code=400,
                content={"success": False, "message": f"Error reading file: {str(e)}"}
            )

        # Delete the temporary file
        os.unlink(temp_file_path)

        # Check if the Excel has the required columns
        required_columns = [
            "customer_name", "phone_no", "date", "quotation_given"
        ]

        # Optional columns
        optional_columns = ["address", "requirements", "quotation_amount"]

        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return JSONResponse(
                status_code=400,
                content={
                    "success": False,
                    "message": f"Missing required columns: {', '.join(missing_columns)}",
                    "required_columns": required_columns
                }
            )

        # Process each row in the Excel file
        success_count = 0
        error_count = 0
        errors = []

        for index, row in df.iterrows():
            try:
                # Parse date
                try:
                    if isinstance(row["date"], str):
                        # Try different date formats
                        date_formats = ["%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"]
                        date_obj = None

                        for fmt in date_formats:
                            try:
                                date_obj = datetime.strptime(row["date"], fmt).date()
                                break
                            except ValueError:
                                continue

                        if date_obj is None:
                            raise ValueError(f"Could not parse date: {row['date']}")
                    else:
                        # If it's already a datetime or date object
                        date_obj = row["date"].date() if hasattr(row["date"], "date") else row["date"]
                except Exception as e:
                    # Default to today if date parsing fails
                    date_obj = datetime.now().date()
                    errors.append(f"Row {index+2}: Date parsing error, using today's date: {str(e)}")

                # Convert quotation_given to boolean
                if isinstance(row["quotation_given"], str):
                    quotation_given = row["quotation_given"].lower() in ["yes", "true", "1", "y"]
                else:
                    quotation_given = bool(row["quotation_given"])

                # Get quotation amount if quotation is given
                quotation_amount = None
                if quotation_given and "quotation_amount" in row and pd.notna(row["quotation_amount"]):
                    try:
                        quotation_amount = float(row["quotation_amount"])
                    except (ValueError, TypeError):
                        quotation_amount = 0.0
                        errors.append(f"Row {index+2}: Invalid quotation amount, using 0.0")

                # Create enquiry data dictionary
                enquiry_data = {
                    "customer_name": str(row["customer_name"]),
                    "phone_no": str(row["phone_no"]),
                    "address": str(row["address"]) if "address" in row and pd.notna(row["address"]) else "",
                    "requirements": str(row["requirements"]) if "requirements" in row and pd.notna(row["requirements"]) else "",
                    "quotation_given": quotation_given,
                    "quotation_amount": quotation_amount if quotation_given else None,
                    "date": date_obj
                }

                # Create new enquiry
                crud.create_enquiry(db=db, enquiry_data=enquiry_data)
                success_count += 1

            except Exception as e:
                error_count += 1
                errors.append(f"Row {index+2}: {str(e)}")
                continue

        # Format errors for display
        formatted_errors = None
        if errors:
            formatted_errors = "<br>".join(errors)

        # Return the results
        return JSONResponse(
            content={
                "success": True,
                "count": success_count,
                "message": f"Successfully imported {success_count} enquiries. {error_count} failed.",
                "errors": formatted_errors if formatted_errors else None
            }
        )
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error importing enquiries from Excel: {e}")
        print(f"Error details: {error_details}")
        return JSONResponse(
            status_code=500,
            content={"success": False, "message": f"Error importing enquiries: {str(e)}"}
        )


@router.get("/api/enquiries/download-template")
async def download_enquiry_template():
    """Download a template Excel file for enquiry import"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import os
        from datetime import datetime

        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enquiry Template"

        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add headers
        headers = [
            "customer_name", "phone_no", "address", "requirements",
            "quotation_given", "quotation_amount", "date"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Add sample data
        sample_data = [
            ["John Doe", "9876543210", "123 Main St, City", "Need AC repair", "Yes", 1500, datetime.now().strftime("%d-%m-%Y")],
            ["Jane Smith", "8765432109", "456 Park Ave, Town", "New AC installation", "No", "", datetime.now().strftime("%d-%m-%Y")]
        ]

        for row_num, data in enumerate(sample_data, 2):
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = thin_border

        # Add instructions in a new sheet
        ws_instructions = wb.create_sheet(title="Instructions")
        instructions = [
            ["Enquiry Import Instructions"],
            [""],
            ["Required Columns:"],
            ["customer_name - Customer's full name (required)"],
            ["phone_no - Customer's phone number (required)"],
            ["date - Date in DD-MM-YYYY format (required)"],
            ["quotation_given - Yes/No or True/False (required)"],
            [""],
            ["Optional Columns:"],
            ["address - Customer's address"],
            ["requirements - Customer's requirements or notes"],
            ["quotation_amount - Amount if quotation is given (required if quotation_given is Yes)"],
            [""],
            ["Notes:"],
            ["1. The system will automatically generate unique enquiry numbers"],
            ["2. For large imports, consider splitting into multiple files of 50-100 enquiries each"],
            ["3. Date formats accepted: DD-MM-YYYY, YYYY-MM-DD, MM/DD/YYYY, DD/MM/YYYY"]
        ]

        for row_num, instruction in enumerate(instructions, 1):
            ws_instructions.cell(row=row_num, column=1).value = instruction[0]
            if row_num == 1:
                ws_instructions.cell(row=row_num, column=1).font = Font(size=14, bold=True)

        # Adjust column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20

        # Create directory for templates if it doesn't exist
        os.makedirs("exports", exist_ok=True)
        template_path = os.path.join("exports", "enquiry_import_template.xlsx")

        # Save the workbook
        wb.save(template_path)

        # Return the template file
        return FileResponse(
            template_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="enquiry_import_template.xlsx"
        )
    except Exception as e:
        print(f"Error creating enquiry template: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@router.get("/download-quotation/{filename}")
async def download_quotation_file(
    request: Request,
    filename: str,
    db: Session = Depends(database.get_db)
):
    """Download a quotation file"""
    try:
        # Get current user from cookie
        user = await get_current_user_from_cookie(request, db)

        # Redirect to login if not authenticated
        if not user:
            return RedirectResponse(url="/login", status_code=303)

        # Construct the file path
        file_path = os.path.join(QUOTATIONS_FOLDER, filename)

        # Check if file exists
        if not os.path.exists(file_path):
            return templates.TemplateResponse(
                "error.html",
                {"request": request, "error_message": "File not found", "user": user},
                status_code=404
            )

        # Return the file
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type="application/octet-stream"
        )
    except Exception as e:
        print(f"Error downloading quotation file: {e}")
        return templates.TemplateResponse(
            "error.html",
            {"request": request, "error_message": f"Error downloading file: {str(e)}", "user": user}
        )

@router.get("/api/enquiries/filtered")
async def api_get_filtered_enquiries(
    customer_name: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    quotation_given: Optional[bool] = None,
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(database.get_db)
):
    """Get filtered enquiries as JSON"""
    filters = {}

    if customer_name:
        filters["customer_name"] = customer_name

    if date_from:
        try:
            filters["date_from"] = datetime.strptime(date_from, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_from format. Use YYYY-MM-DD")

    if date_to:
        try:
            filters["date_to"] = datetime.strptime(date_to, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date_to format. Use YYYY-MM-DD")

    if quotation_given is not None:
        filters["quotation_given"] = quotation_given

    enquiries = crud.get_filtered_enquiries(db, filters, limit, offset)
    total = crud.get_filtered_enquiries_count(db, filters)

    result = []
    for e in enquiries:
        result.append({
            "id": e.id,
            "enquiry_number": e.enquiry_number,
            "date": e.date.isoformat(),
            "customer_name": e.customer_name,
            "phone_no": e.phone_no,
            "address": e.address,
            "requirements": e.requirements,
            "quotation_given": e.quotation_given,
            "quotation_amount": e.quotation_amount,
            "created_at": e.created_at.isoformat(),
            "updated_at": e.updated_at.isoformat()
        })

    return {"total": total, "items": result}


@router.get("/enquiries/export")
async def export_enquiries(
    request: Request,
    format: str = "csv",
    db: Session = Depends(database.get_db)
):
    """Export all enquiries to CSV or JSON"""
    try:
        # Get current user from cookie
        user = await get_current_user_from_cookie(request, db)

        # Redirect to login if not authenticated
        if not user:
            return RedirectResponse(url="/login?next=/enquiries/export", status_code=303)

        # Get all enquiries
        enquiries = crud.get_all_enquiries(db)

        if format.lower() == "json":
            # Convert to JSON
            result = []
            for e in enquiries:
                result.append({
                    "id": e.id,
                    "enquiry_number": e.enquiry_number,
                    "date": e.date.isoformat(),
                    "customer_name": e.customer_name,
                    "phone_no": e.phone_no,
                    "address": e.address,
                    "requirements": e.requirements,
                    "quotation_given": e.quotation_given,
                    "quotation_amount": e.quotation_amount,
                    "created_at": e.created_at.isoformat(),
                    "updated_at": e.updated_at.isoformat()
                })

            return JSONResponse(content=result)
        else:
            # Default to CSV
            output = StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow([
                "Enquiry Number", "Date", "Customer Name", "Phone",
                "Address", "Requirements", "Quotation Given", "Quotation Amount"
            ])

            # Write data
            for e in enquiries:
                writer.writerow([
                    e.enquiry_number,
                    e.date.isoformat(),
                    e.customer_name,
                    e.phone_no,
                    e.address,
                    e.requirements,
                    "Yes" if e.quotation_given else "No",
                    e.quotation_amount if e.quotation_amount else ""
                ])

            # Return CSV file
            output.seek(0)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=enquiries_export_{datetime.now().strftime('%Y%m%d')}.csv"}
            )
    except Exception as e:
        print(f"Error exporting enquiries: {e}")
        return templates.TemplateResponse(
            "error.html",
            {"request": request, "error_message": f"Error exporting enquiries: {str(e)}"}
        )





