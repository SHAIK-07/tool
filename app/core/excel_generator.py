"""
Excel generation utilities for exporting data to Excel files.
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from sqlalchemy.orm import Session
from app.db import models


def export_services_to_excel(services):
    """
    Export service records to Excel file

    Args:
        services: List of service records

    Returns:
        Tuple of (filename, file_path)
    """
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Services"

        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add headers
        headers = [
            "Service Code", "Date", "Service Name", "Employee Name",
            "Description", "Price (₹)", "GST Rate (%)", "Total Price (₹)"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Add data rows
        row_num = 2
        for service in services:
            # Calculate total price with GST
            price = service.price
            gst_rate = service.gst_rate if hasattr(service, 'gst_rate') and service.gst_rate is not None else 18.0
            total_price = price + (price * gst_rate / 100)

            # Format date
            date_str = service.date.strftime('%d-%m-%Y') if hasattr(service.date, 'strftime') else str(service.date)

            # Add service data
            ws.cell(row=row_num, column=1).value = service.service_code
            ws.cell(row=row_num, column=2).value = date_str
            ws.cell(row=row_num, column=3).value = service.service_name
            ws.cell(row=row_num, column=4).value = service.employee_name
            ws.cell(row=row_num, column=5).value = service.description
            ws.cell(row=row_num, column=6).value = price
            ws.cell(row=row_num, column=7).value = gst_rate
            ws.cell(row=row_num, column=8).value = round(total_price, 2)

            # Apply borders to all cells in the row
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).border = thin_border

                # Right-align numeric columns
                if col_num in [6, 7, 8]:
                    ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='right')

            row_num += 1

        # Auto-adjust column widths
        for col_num, _ in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 15

        # Make the description column wider
        ws.column_dimensions['E'].width = 40

        # Generate filename with date
        current_date = datetime.now().strftime('%Y%m%d')
        filename = f"Sunmax_Services_{current_date}.xlsx"

        # Create directory for exports if it doesn't exist
        os.makedirs("exports", exist_ok=True)
        file_path = os.path.join("exports", filename)

        # Save the workbook to a file
        wb.save(file_path)

        return filename, file_path
    except Exception as e:
        print(f"Error exporting services to Excel: {e}")
        raise


def export_inventory_to_excel(items):
    """
    Export inventory items to Excel file

    Args:
        items: List of inventory items

    Returns:
        Tuple of (filename, file_path)
    """
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory"

        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add headers
        headers = [
            "Item Code", "Item Name", "HSN Code", "Purchase Price (₹)",
            "Margin (%)", "Selling Price (₹)", "GST Rate (%)", "Quantity",
            "Unit", "Supplier Name", "Supplier GST", "Date Added"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Add data rows
        row_num = 2
        for item in items:
            # Calculate selling price
            purchase_price = item.purchase_price_per_unit
            margin = item.margin if hasattr(item, 'margin') and item.margin is not None else 20
            selling_price = purchase_price * (1 + (margin / 100))

            # Format date
            date_str = item.date.strftime('%d-%m-%Y') if hasattr(item.date, 'strftime') else str(item.date)

            # Add item data
            ws.cell(row=row_num, column=1).value = item.item_code
            ws.cell(row=row_num, column=2).value = item.item_name
            ws.cell(row=row_num, column=3).value = item.hsn_code
            ws.cell(row=row_num, column=4).value = item.purchase_price_per_unit
            ws.cell(row=row_num, column=5).value = margin
            ws.cell(row=row_num, column=6).value = round(selling_price, 2)
            ws.cell(row=row_num, column=7).value = item.gst_rate
            ws.cell(row=row_num, column=8).value = item.quantity
            ws.cell(row=row_num, column=9).value = item.unit_of_measurement
            ws.cell(row=row_num, column=10).value = item.supplier_name
            ws.cell(row=row_num, column=11).value = item.supplier_gst_number
            ws.cell(row=row_num, column=12).value = date_str

            # Apply borders to all cells in the row
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).border = thin_border

                # Right-align numeric columns
                if col_num in [4, 5, 6, 7, 8]:
                    ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='right')

            row_num += 1

        # Auto-adjust column widths
        for col_num, _ in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 15

        # Make the item name column wider
        ws.column_dimensions['B'].width = 30

        # Create an import template sheet
        template_ws = wb.create_sheet(title="Import Template")
        template_headers = [
            "item_code", "item_name", "hsn_code", "purchase_price_per_unit", "margin",
            "gst_rate", "quantity", "unit_of_measurement", "supplier_name",
            "supplier_gst_number"
        ]

        # Add headers to template worksheet
        for col_num, header in enumerate(template_headers, 1):
            cell = template_ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

        # Add note about item_code
        note_cell = template_ws.cell(row=2, column=1)
        note_cell.value = "Optional - leave blank for new items, provide for updating existing items"
        note_cell.font = Font(italic=True, color="FF0000")
        template_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

        # Add sample data row
        sample_data = [
            "", "Sample Item", "12345678", 1000.00, 20.0,
            18.0, 10, "No.s", "Supplier Name", "GSTIN12345"
        ]

        for col_num, value in enumerate(sample_data, 1):
            cell = template_ws.cell(row=3, column=col_num)
            cell.value = value

        # Adjust column widths
        for col in range(1, len(template_headers) + 1):
            template_ws.column_dimensions[get_column_letter(col)].width = 20

        # Generate filename with date
        current_date = datetime.now().strftime('%Y%m%d')
        filename = f"Sunmax_Inventory_{current_date}.xlsx"

        # Create directory for exports if it doesn't exist
        os.makedirs("exports", exist_ok=True)
        file_path = os.path.join("exports", filename)

        # Save the workbook to a file
        wb.save(file_path)

        return filename, file_path
    except Exception as e:
        print(f"Error exporting inventory to Excel: {e}")
        raise


def export_invoices_to_excel(invoices, db: Session, payment_status=None):
    """
    Export invoices to Excel file with detailed information for GST filing

    Args:
        invoices: List of invoice dictionaries
        db: Database session
        payment_status: Optional filter for payment status

    Returns:
        Tuple of (filename, file_path)
    """
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoices"

        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Create a single sheet with all invoice and item details
        ws.title = "Invoice Data"

        # Add headers - Updated format with requested columns
        headers = [
            "HSN Code", "Invoice Number", "Date", "Customer Name", "Customer GST",
            "Payment Method", "Payment Status", "Item Code", "Item Name", "Quantity",
            "Subtotal", "Discount", "Taxable Amount", "GST", "Total"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Add data - Each item gets its own row
        row_num = 2
        for invoice in invoices:

            # Get invoice items directly from the database for more reliable data
            invoice_items_db = db.query(models.InvoiceItem).join(
                models.Invoice, models.Invoice.id == models.InvoiceItem.invoice_id
            ).filter(
                models.Invoice.invoice_number == invoice["invoice_number"]
            ).all()

            # If no items, add one row for the invoice
            if not invoice_items_db:
                # Add invoice data with HSN Code first
                ws.cell(row=row_num, column=1).value = "N/A"  # HSN Code (N/A for no items)
                ws.cell(row=row_num, column=2).value = invoice["invoice_number"]
                ws.cell(row=row_num, column=3).value = invoice["date"].strftime('%d-%m-%Y') if hasattr(invoice["date"], 'strftime') else str(invoice["date"])
                ws.cell(row=row_num, column=4).value = invoice["customer_name"]
                ws.cell(row=row_num, column=5).value = invoice["customer_gst"]
                ws.cell(row=row_num, column=6).value = invoice["payment_method"]
                ws.cell(row=row_num, column=7).value = invoice["payment_status"]
                ws.cell(row=row_num, column=8).value = "No items"  # Item Code
                ws.cell(row=row_num, column=9).value = "No items"  # Item Name
                ws.cell(row=row_num, column=10).value = 0  # Quantity
                ws.cell(row=row_num, column=11).value = f"Rs.{invoice['subtotal']:.1f}"  # Subtotal
                ws.cell(row=row_num, column=12).value = f"Rs.{invoice.get('discount', 0):.1f}"  # Discount

                # Calculate taxable amount (subtotal - discount)
                taxable_amount = invoice["subtotal"] - invoice.get("discount", 0)
                ws.cell(row=row_num, column=13).value = f"Rs.{taxable_amount:.1f}"  # Taxable Amount

                ws.cell(row=row_num, column=14).value = f"Rs.{invoice['total_gst']:.1f}"  # GST
                ws.cell(row=row_num, column=15).value = f"Rs.{invoice['total_amount']:.1f}"  # Total

                # Apply borders to all cells
                for col_num in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col_num).border = thin_border

                row_num += 1
            else:
                # Add a row for each item
                for item in invoice_items_db:
                    # Add invoice data with HSN Code first (repeated for each item)
                    ws.cell(row=row_num, column=1).value = item.hsn_code  # HSN Code first
                    ws.cell(row=row_num, column=2).value = invoice["invoice_number"]
                    ws.cell(row=row_num, column=3).value = invoice["date"].strftime('%d-%m-%Y') if hasattr(invoice["date"], 'strftime') else str(invoice["date"])
                    ws.cell(row=row_num, column=4).value = invoice["customer_name"]
                    ws.cell(row=row_num, column=5).value = invoice["customer_gst"]
                    ws.cell(row=row_num, column=6).value = invoice["payment_method"]
                    ws.cell(row=row_num, column=7).value = invoice["payment_status"]

                    # Add item data
                    ws.cell(row=row_num, column=8).value = item.item_code
                    ws.cell(row=row_num, column=9).value = item.item_name
                    ws.cell(row=row_num, column=10).value = item.quantity

                    # Add calculation fields
                    item_subtotal = item.price * item.quantity
                    ws.cell(row=row_num, column=11).value = f"Rs.{item_subtotal:.1f}"  # Subtotal

                    # Get discount - check multiple possible fields
                    discount = getattr(item, 'discount', 0)
                    if discount == 0:
                        discount_percent = getattr(item, 'discount_percent', 0)
                        if discount_percent > 0:
                            discount = item_subtotal * (discount_percent / 100)
                        else:
                            discount_amount = getattr(item, 'discount_amount', 0)
                            if discount_amount > 0:
                                discount = discount_amount

                    ws.cell(row=row_num, column=12).value = f"Rs.{discount:.1f}"  # Discount

                    # Calculate taxable amount
                    taxable_amount = item_subtotal - discount
                    ws.cell(row=row_num, column=13).value = f"Rs.{taxable_amount:.1f}"  # Taxable Amount

                    # GST amount
                    gst_amount = item.gst_amount if hasattr(item, 'gst_amount') else (taxable_amount * (item.gst_rate / 100))
                    ws.cell(row=row_num, column=14).value = f"Rs.{gst_amount:.1f}"  # GST

                    # Total
                    total = taxable_amount + gst_amount
                    ws.cell(row=row_num, column=15).value = f"Rs.{total:.1f}"  # Total

                    # Apply borders to all cells
                    for col_num in range(1, len(headers) + 1):
                        ws.cell(row=row_num, column=col_num).border = thin_border

                    row_num += 1

            # No individual invoice sheets - all data is in the main sheet

        # Auto-adjust column widths for the single sheet
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Generate filename with date
        current_date = datetime.now().strftime('%Y%m%d')
        status_suffix = f"_{payment_status}" if payment_status and payment_status != "all" else ""
        filename = f"Sunmax_Invoices{status_suffix}_{current_date}.xlsx"

        # Create directory for exports if it doesn't exist
        os.makedirs("exports", exist_ok=True)
        file_path = os.path.join("exports", filename)

        # Save the workbook to a file
        wb.save(file_path)

        return filename, file_path
    except Exception as e:
        print(f"Error exporting invoices to Excel: {e}")
        raise


def export_customers_to_excel(customers, status_filter=None):
    """
    Export customer records to Excel file

    Args:
        customers: List of customer records
        status_filter: Optional filter for payment status

    Returns:
        Tuple of (filename, file_path)
    """
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customers"

        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add headers
        headers = [
            "Customer Code", "Date", "Customer Name", "Phone Number", "Address",
            "Product Description", "Payment Method", "Payment Status",
            "Total Amount (₹)", "Amount Paid (₹)", "Balance Due (₹)"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Add data rows
        row_num = 2
        for customer in customers:
            # Calculate balance due
            total_amount = customer.total_amount if hasattr(customer, 'total_amount') else 0
            amount_paid = customer.amount_paid if hasattr(customer, 'amount_paid') else 0
            balance_due = total_amount - amount_paid

            # Format date
            date_str = customer.date.strftime('%d-%m-%Y') if hasattr(customer.date, 'strftime') else str(customer.date)

            # Add customer data
            ws.cell(row=row_num, column=1).value = customer.customer_code
            ws.cell(row=row_num, column=2).value = date_str
            ws.cell(row=row_num, column=3).value = customer.customer_name
            ws.cell(row=row_num, column=4).value = customer.phone_no
            ws.cell(row=row_num, column=5).value = customer.address if hasattr(customer, 'address') else ""
            ws.cell(row=row_num, column=6).value = customer.product_description if hasattr(customer, 'product_description') else ""
            ws.cell(row=row_num, column=7).value = customer.payment_method if hasattr(customer, 'payment_method') else ""
            ws.cell(row=row_num, column=8).value = customer.payment_status
            ws.cell(row=row_num, column=9).value = total_amount
            ws.cell(row=row_num, column=10).value = amount_paid
            ws.cell(row=row_num, column=11).value = balance_due

            # Apply borders to all cells in the row
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).border = thin_border

                # Right-align numeric columns
                if col_num in [9, 10, 11]:
                    ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='right')

            row_num += 1

        # Auto-adjust column widths
        for col_num, _ in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 15

        # Make the address and product description columns wider
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30

        # Generate filename with date
        current_date = datetime.now().strftime('%Y%m%d')
        status_suffix = f"_{status_filter}" if status_filter and status_filter != "all" else ""
        filename = f"Sunmax_Customers{status_suffix}_{current_date}.xlsx"

        # Create directory for exports if it doesn't exist
        os.makedirs("exports", exist_ok=True)
        file_path = os.path.join("exports", filename)

        # Save the workbook to a file
        wb.save(file_path)

        return filename, file_path
    except Exception as e:
        print(f"Error exporting customers to Excel: {e}")
        raise


def export_quotations_to_excel(quotations):
    """
    Export quotations to Excel with formatting

    Args:
        quotations: List of quotation objects

    Returns:
        tuple: (filename, file_path)
    """
    import pandas as pd
    import os
    from datetime import datetime

    # Create a DataFrame from the quotations
    data = []
    for q in quotations:
        data.append({
            "Quote Number": q.quote_number,
            "Date": q.date.strftime("%d-%m-%Y"),
            "Customer Name": q.customer_name,
            "Customer Phone": q.customer_phone,
            "Customer Email": q.customer_email or "",
            "Customer Address": q.customer_address or "",
            "Asked About": q.asked_about or "",
            "Subtotal": q.subtotal,
            "GST": q.total_gst,
            "Total Amount": q.total_amount,
            "Created At": q.created_at.strftime("%d-%m-%Y %H:%M:%S")
        })

    # Create DataFrame
    df = pd.DataFrame(data)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"quotations_{timestamp}.xlsx"

    # Create directory if it doesn't exist
    export_dir = os.path.join("static", "exports")
    os.makedirs(export_dir, exist_ok=True)

    # Full path to the file
    file_path = os.path.join(export_dir, filename)

    # Create Excel file with formatting
    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name="Quotations", index=False)

        # Get the xlsxwriter workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets["Quotations"]

        # Add formatting
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4285F4',
            'color': 'white',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })

        # Currency format
        currency_format = workbook.add_format({
            'num_format': '₹#,##0.00',
            'align': 'right'
        })

        # Date format
        date_format = workbook.add_format({
            'num_format': 'dd-mm-yyyy',
            'align': 'center'
        })

        # Apply header format
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)

        # Set column widths and formats
        worksheet.set_column('A:A', 15)  # Quote Number
        worksheet.set_column('B:B', 12)  # Date
        worksheet.set_column('C:C', 25)  # Customer Name
        worksheet.set_column('D:D', 15)  # Customer Phone
        worksheet.set_column('E:E', 25)  # Customer Email
        worksheet.set_column('F:F', 30)  # Customer Address
        worksheet.set_column('G:G', 30)  # Asked About
        worksheet.set_column('H:H', 15, currency_format)  # Subtotal
        worksheet.set_column('I:I', 15, currency_format)  # GST
        worksheet.set_column('J:J', 15, currency_format)  # Total Amount
        worksheet.set_column('K:K', 20)  # Created At

        # Add auto-filter
        worksheet.autofilter(0, 0, len(df), len(df.columns) - 1)

    return filename, file_path


def export_enquiries_to_excel(enquiries):
    """
    Export enquiries to Excel with formatting

    Args:
        enquiries: List of enquiry objects

    Returns:
        tuple: (filename, file_path)
    """
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enquiries"

        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add headers
        headers = [
            "Enquiry Number", "Date", "Customer Name", "Phone Number",
            "Address", "Requirements", "Quotation Given", "Quotation Amount (₹)"
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Add data rows
        row_num = 2
        for enquiry in enquiries:
            # Format date
            date_str = enquiry.date.strftime('%d-%m-%Y') if hasattr(enquiry.date, 'strftime') else str(enquiry.date)

            # Add enquiry data
            ws.cell(row=row_num, column=1).value = enquiry.enquiry_number
            ws.cell(row=row_num, column=2).value = date_str
            ws.cell(row=row_num, column=3).value = enquiry.customer_name
            ws.cell(row=row_num, column=4).value = enquiry.phone_no
            ws.cell(row=row_num, column=5).value = enquiry.address
            ws.cell(row=row_num, column=6).value = enquiry.requirements
            ws.cell(row=row_num, column=7).value = "Yes" if enquiry.quotation_given else "No"
            ws.cell(row=row_num, column=8).value = enquiry.quotation_amount if enquiry.quotation_given else "-"

            # Apply borders to all cells in the row
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).border = thin_border

                # Right-align numeric columns
                if col_num == 8:
                    ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='right')

            row_num += 1

        # Auto-adjust column widths
        for col_num, _ in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 15

        # Make the requirements column wider
        ws.column_dimensions['F'].width = 40
        # Make the address column wider
        ws.column_dimensions['E'].width = 30

        # Generate filename with date
        current_date = datetime.now().strftime('%Y%m%d')
        filename = f"Sunmax_Enquiries_{current_date}.xlsx"

        # Create directory for exports if it doesn't exist
        os.makedirs("exports", exist_ok=True)
        file_path = os.path.join("exports", filename)

        # Save the workbook to a file
        wb.save(file_path)

        return filename, file_path
    except Exception as e:
        print(f"Error exporting enquiries to Excel: {e}")
        raise
